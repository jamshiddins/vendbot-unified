from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import logging

from ..models.audit_models import ReconciliationResult, Discrepancy, DiscrepancyType

logger = logging.getLogger(__name__)

class ExcelReporter:
    """Генератор Excel отчетов"""
    
    def __init__(self):
        self.wb = None
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """Создание стилей для отчета"""
        return {
            'header': {
                'font': Font(bold=True, size=14),
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                'font_color': Font(color="FFFFFF", bold=True),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'subheader': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
                'alignment': Alignment(horizontal="left", vertical="center")
            },
            'critical': {
                'fill': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            },
            'high': {
                'fill': PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
            },
            'medium': {
                'fill': PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
            },
            'low': {
                'fill': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def generate_reconciliation_report(self, 
                                     result: ReconciliationResult,
                                     output_path: Path) -> Path:
        """Генерация отчета о сверке"""
        logger.info(f"Generating reconciliation report to {output_path}")
        
        self.wb = Workbook()
        
        # Удаляем дефолтный лист
        self.wb.remove(self.wb.active)
        
        # Создаем листы
        self._create_summary_sheet(result)
        self._create_discrepancies_sheet(result.discrepancies)
        self._create_machine_summary_sheet(result.summary_by_machine)
        self._create_payment_summary_sheet(result.summary_by_payment)
        
        # Сохраняем файл
        self.wb.save(output_path)
        logger.info(f"Report saved to {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, result: ReconciliationResult):
        """Создание листа со сводкой"""
        ws = self.wb.create_sheet("Сводка")
        
        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Отчет о сверке за период {result.period_start.date()} - {result.period_end.date()}"
        ws['A1'].font = self.styles['header']['font']
        ws['A1'].fill = self.styles['header']['fill']
        ws['A1'].alignment = self.styles['header']['alignment']
        
        # Основные показатели
        data = [
            ["Показатель", "Значение"],
            ["Всего продаж", result.total_sales],
            ["Всего чеков", result.total_receipts],
            ["Всего QR транзакций", result.total_transactions],
            ["Сверено успешно", result.matched_count],
            ["Выявлено несоответствий", len(result.discrepancies)],
            ["% успешной сверки", f"{(result.matched_count / result.total_sales * 100):.1f}%" if result.total_sales > 0 else "0%"]
        ]
        
        start_row = 3
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                if row_idx == 0:
                    cell.font = Font(bold=True)
                    cell.fill = self.styles['subheader']['fill']
                cell.border = self.styles['border']
        
        # Автоширина колонок
        self._autosize_columns(ws)
    
    def _create_discrepancies_sheet(self, discrepancies: List[Discrepancy]):
        """Создание листа с несоответствиями"""
        ws = self.wb.create_sheet("Несоответствия")
        
        # Заголовки
        headers = ["Тип", "Машина", "Дата/Время", "Описание", "Сумма разницы", "Важность"]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=1, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.styles['subheader']['fill']
            cell.border = self.styles['border']
        
        # Данные
        for row_idx, discrepancy in enumerate(discrepancies, start=2):
            ws.cell(row=row_idx, column=1, value=discrepancy.type.value)
            ws.cell(row=row_idx, column=2, value=discrepancy.machine_id)
            ws.cell(row=row_idx, column=3, value=discrepancy.datetime.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=4, value=discrepancy.description)
            ws.cell(row=row_idx, column=5, value=float(discrepancy.amount_difference) if discrepancy.amount_difference else "")
            
            severity_cell = ws.cell(row=row_idx, column=6, value=discrepancy.severity)
            severity_cell.fill = self.styles.get(discrepancy.severity, {}).get('fill', PatternFill())
            
            # Применяем границы
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = self.styles['border']
        
        self._autosize_columns(ws)
    
    def _create_machine_summary_sheet(self, summary: Dict[str, Dict[str, Any]]):
        """Создание листа со сводкой по машинам"""
        ws = self.wb.create_sheet("По машинам")
        
        # Заголовки
        headers = ["Машина", "Продаж", "Сумма", "Несоответствий", "Наличные", "Карта", "QR"]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=1, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.styles['subheader']['fill']
            cell.border = self.styles['border']
        
        # Данные
        row_idx = 2
        for machine_id, data in sorted(summary.items()):
            ws.cell(row=row_idx, column=1, value=machine_id)
            ws.cell(row=row_idx, column=2, value=data['total_sales'])
            ws.cell(row=row_idx, column=3, value=float(data['total_amount']))
            ws.cell(row=row_idx, column=4, value=data['discrepancies'])
            
            # Методы оплаты
            ws.cell(row=row_idx, column=5, value=data['payment_methods'].get('cash', 0))
            ws.cell(row=row_idx, column=6, value=data['payment_methods'].get('card', 0))
            
            qr_total = (data['payment_methods'].get('qr_click', 0) +
                       data['payment_methods'].get('qr_payme', 0) +
                       data['payment_methods'].get('qr_uzum', 0))
            ws.cell(row=row_idx, column=7, value=qr_total)
            
            # Применяем границы
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = self.styles['border']
            
            row_idx += 1
        
        self._autosize_columns(ws)
    
    def _create_payment_summary_sheet(self, summary: Dict[str, Dict[str, Any]]):
        """Создание листа со сводкой по методам оплаты"""
        ws = self.wb.create_sheet("По оплате")
        
        # Заголовки
        headers = ["Метод оплаты", "Количество", "Сумма", "Сверено", "Не сверено", "% успеха"]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=1, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.styles['subheader']['fill']
            cell.border = self.styles['border']
        
        # Данные
        row_idx = 2
        for payment_method, data in sorted(summary.items()):
            ws.cell(row=row_idx, column=1, value=payment_method)
            ws.cell(row=row_idx, column=2, value=data['count'])
            ws.cell(row=row_idx, column=3, value=float(data['amount']))
            ws.cell(row=row_idx, column=4, value=data['matched'])
            ws.cell(row=row_idx, column=5, value=data['unmatched'])
            
            success_rate = (data['matched'] / data['count'] * 100) if data['count'] > 0 else 0
            ws.cell(row=row_idx, column=6, value=f"{success_rate:.1f}%")
            
            # Подсветка проблемных методов
            if success_rate < 90:
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = self.styles['medium']['fill']
            
            # Применяем границы
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = self.styles['border']
            
            row_idx += 1
        
        self._autosize_columns(ws)
    
    def _autosize_columns(self, ws):
        """Автоматическая настройка ширины колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width